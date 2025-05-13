# -*- coding: utf-8 -*-
"""
Created on Sun Jul 21 01:17:29 2024

@author: anash
"""

import os
import cv2
import openpyxl
import torch
import lightning as L
import numpy as np
import random as rndm
import csv
import pickle
from os.path import isfile, join
from CNNInference import CNNInference
from LSTMClassifier import LSTMClassifierTrain
from torch.utils.data import TensorDataset, DataLoader
from torch.nn.utils.rnn import pack_sequence, pad_packed_sequence


FIRST_RUN = True
CREATE_NEW_DS = True

image_model = r".\runs\segment\train23\weights\best.pt"
path = r"C:\Users\anash\OneDrive\Dokumente\datasets_eval"
datasets = [dataset for dataset in os.listdir(path) if "Dataset " in dataset]
evaluated_datasets =  [evaluated for evaluated in os.listdir(path) if "dataset" in evaluated]
checkpoint = r"C:\Users\anash\Documents\lightning_logs\version_355\checkpoints\epoch=199-step=48000.ckpt"

def createDatasets(labels, features, train_size):
    stable_clips = [clip for clip, label in labels if label == 0]
    metastable_clips = [clip for clip, label in labels if label == 1]
    unstable_clips = [clip for clip, label in labels if label == 2]
    
    stable = [[clip, features] for clip, features in features if clip in stable_clips]
    metastable = [[clip, features] for clip, features in features if clip in metastable_clips]
    unstable = [[clip, features] for clip, features in features if clip in unstable_clips]
    
    n = int(train_size/3)
    
    if len(stable) < n or len(metastable) < n or len(unstable) < n:
        n = min(len(stable), len(metastable), len(unstable))
    
    rndm_stable = rndm.sample(stable, n)
    rndm_metastable = rndm.sample(metastable, n)
    rndm_unstable = rndm.sample(unstable, n)
    
    train_features = []
    train_clips = []
    train_labels = []
    for i in range(n):
        # Eine zufällige Reihenfolge für die Elemente bestimmen
        elements = [(rndm_stable[i], 0), (rndm_metastable[i], 1), (rndm_unstable[i], 2)]
        rndm.shuffle(elements)
        for (clip, feature), label in elements:
            train_clips.append(clip)
            train_features.append(feature)
            train_labels.append(label)
    
    remaining_stable = [elem for elem in stable if elem not in rndm_stable]
    remaining_metastable = [elem for elem in metastable if elem not in rndm_metastable]
    remaining_unstable = [elem for elem in unstable if elem not in rndm_unstable]
    
    val = remaining_stable + remaining_metastable + remaining_unstable
    val_clips = [clip for clip, feature in val]
    val_features = [feature for clip, feature in val] 
    val_labels = ([0] * len(remaining_stable)) + ([1] * len(remaining_metastable)) + ([2] * len(remaining_unstable))
    
    train_dataset = {
        'clips': train_clips,
        'features': train_features,
        'labels': train_labels
    }
    
    val_dataset = {
        'clips': val_clips,
        'features': val_features,
        'labels': val_labels
    }
    
    return train_dataset, val_dataset


if __name__=="__main__":
    
    if FIRST_RUN:
        sequences = list()
        labels = list()
        
        for ds_no in range(9):
            wb = openpyxl.load_workbook(join(path,evaluated_datasets[ds_no]))
            ws = wb.active
            clips = []
            label = []
            for i in range(2,ws.max_row+1):
                if ws.cell(row=i, column=3).value < 3:
                    clips.append(f"clip{int(ws.cell(row=i, column=1).value.strip("Clip ")):03}.mp4")
                    label.append([int(ws.cell(row=i, column=1).value.strip("Clip ")), ws.cell(row=i, column=2).value])
            labels = labels + label
       
            CNNmodel = CNNInference(image_model)
            
            for clip in clips:
                CNNmodel.detect(join(path, datasets[ds_no], clip), conf=0.001)
                
                in1 = torch.tensor(CNNmodel.dis_x_list, dtype=torch.float32).view(-1, 1)
                in2 = torch.tensor(CNNmodel.dis_y_list, dtype=torch.float32).view(-1, 1)
                in3 = torch.tensor(CNNmodel.la_len_list, dtype=torch.float32).view(-1, 1)
                in4 = torch.tensor(CNNmodel.ua_len_list, dtype=torch.float32).view(-1, 1)
                
                inputs = torch.cat([in1, in2, in3, in4], dim=1)
                sequences.append([int(clip.split(".")[0].strip("clip")), inputs])
                
        torch.save(sequences, "var_detected_features.pt")
        torch.save(labels, "var_labels.pt")


    if CREATE_NEW_DS:
        loaded_features = torch.load(r"var_detected_features.pt")
        loaded_labels = torch.load("var_labels.pt")
    
        train, val = createDatasets(loaded_labels, loaded_features, 300)
        if True:
            with open('train_data.pkl', 'wb') as fp:
                pickle.dump(train, fp)
                print('train data saved successfully to file')
            with open('val_data.pkl', 'wb') as fp:
                pickle.dump(val, fp)
                print('val data saved successfully to file')

    train_labels = torch.tensor(train["labels"])
    val_labels = torch.tensor(val["labels"])
    
    train_packed_sequences = pack_sequence(train["features"],enforce_sorted=False)
    train_unpacked_sequences, lengths = pad_packed_sequence(train_packed_sequences, batch_first=True)
    val_packed_sequences = pack_sequence(val["features"],enforce_sorted=False)
    val_unpacked_sequences, lengths = pad_packed_sequence(val_packed_sequences, batch_first=True)
    
    train_dataset = TensorDataset(train_unpacked_sequences, train_labels) 
    train_dataloader = DataLoader(train_dataset, batch_size=2, shuffle=True)
    val_dataset = TensorDataset(val_unpacked_sequences, val_labels) 
    val_dataloader = DataLoader(val_dataset, batch_size=2, shuffle=False)
    
    model = LSTMClassifierTrain(input_size=4, hidden_size=150, num_classes=3, num_layers=2)
    trainer = L.Trainer(max_epochs=300, log_every_n_steps=50)
    trainer.fit(model, train_dataloaders=train_dataloader, val_dataloaders=val_dataloader)

    model.to("cuda")
    count = 0
    count_no_1s = 0
    accuracy = 0
    inaccuracy = 0
    print("\nAfter training:")
    for idx, input1 in enumerate(val_unpacked_sequences):
        with torch.no_grad():
            prediction = model(input1.unsqueeze(0).to("cuda"))  # Füge Batch-Dimension hinzu
            observed_class = val_labels[idx].item()
            predicted_class = prediction.argmax(dim=1).item()
            if observed_class != 1:
                print(f"Clip {val["clips"][idx]:03}: Observed = {observed_class}, Predicted = {predicted_class}")
                if observed_class == predicted_class:
                    accuracy += 1
                elif np.abs(observed_class-predicted_class) == 2:
                    inaccuracy += 1
                if observed_class != 1:
                    count_no_1s +=1
                count += 1
    print(f"\nRaw prediction accuracy: {round((accuracy/count)*100, 2)} %")
    print(f"Very inaccurate predictions: {round((inaccuracy/count_no_1s)*100, 2)} %")